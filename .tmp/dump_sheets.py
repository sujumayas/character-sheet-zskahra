"""Dump each sheet of the workbook to a JSON file with cell values and formulas.

For each sheet we write .tmp/sheets/<safe-name>.json with:
{
  "name": "<sheet name>",
  "max_row": int,
  "max_col": int,
  "merged": ["A1:C1", ...],
  "cells": [
    {"addr": "A1", "row": 1, "col": 1, "v": <displayed value>, "f": "<formula or null>"},
    ...
  ]
}
We omit cells that are entirely empty (no value AND no formula).
"""

from __future__ import annotations
import json, re, pathlib
from openpyxl import load_workbook

ROOT = pathlib.Path(__file__).parent
SRC = ROOT / "sheet.xlsx"
OUT = ROOT / "sheets"
OUT.mkdir(parents=True, exist_ok=True)

# Two passes: data_only=False keeps formulas, data_only=True returns last-cached values.
wb_f = load_workbook(SRC, data_only=False)
wb_v = load_workbook(SRC, data_only=True)


def safe(name: str) -> str:
    return re.sub(r"[^A-Za-z0-9._-]+", "_", name).strip("_")


index = []
for name in wb_f.sheetnames:
    ws_f = wb_f[name]
    ws_v = wb_v[name]
    cells = []
    for row in ws_f.iter_rows():
        for cell in row:
            f = cell.value if isinstance(cell.value, str) and cell.value.startswith("=") else None
            v = ws_v[cell.coordinate].value
            if f is None and (v is None or v == ""):
                continue
            cells.append({
                "addr": cell.coordinate,
                "row": cell.row,
                "col": cell.column,
                "v": v,
                "f": f,
            })
    merged = [str(r) for r in ws_f.merged_cells.ranges]
    out_name = safe(name)
    out_path = OUT / f"{out_name}.json"
    payload = {
        "name": name,
        "max_row": ws_f.max_row,
        "max_col": ws_f.max_column,
        "merged": merged,
        "cell_count": len(cells),
        "cells": cells,
    }
    out_path.write_text(json.dumps(payload, ensure_ascii=False, default=str))
    size_kb = out_path.stat().st_size / 1024
    index.append({
        "tab": name,
        "file": str(out_path.relative_to(ROOT)),
        "non_empty_cells": len(cells),
        "size_kb": round(size_kb, 1),
        "max_row": ws_f.max_row,
        "max_col": ws_f.max_column,
    })

(ROOT / "sheets" / "_index.json").write_text(json.dumps(index, ensure_ascii=False, indent=2))
for row in index:
    print(f"  {row['tab']:<30} cells={row['non_empty_cells']:>6}  {row['size_kb']:>7.1f} KB  ({row['file']})")
